import string
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

class ExcelUtility:  
    
    def create_excel_file(download_path, file_name):
        book = Workbook()
        new_book = book.active
        new_book.title = "Sök resultat"
        new_book['A1'] = 'Länkar'
        new_book['A1'].font = Font(bold=True)
        new_book.column_dimensions['B'].width = 40
        new_book.column_dimensions['C'].width = 40
        new_book.column_dimensions['D'].width = 40
        book.save(filename = f'{download_path}\{file_name}')
        book.close()
    
    def write_links(download_path, file_name, links):
        book = load_workbook(f'{download_path}\{file_name}')
        ws = book.active
        for i, link in enumerate(links):
            ws.column_dimensions['A'].width = len(link) / 2
            ws['A' + str(i+2)] = links[i]
            ws['A' + str(i+2)].hyperlink = links[i]
            ws['A' + str(i+2)].style = "Hyperlink"
        
        book.save(f'{download_path}\{file_name}')
        book.close()
       
    
    def write_keywords(download_path, file_name, keywords):
        book = load_workbook(f'{download_path}\{file_name}')
        ws = book.active
        for i, keyword in enumerate(keywords):
            if len(keyword) == 0:
                ws[string.ascii_uppercase[i+1] + str(1)] = 'Inget nyckelord'
                ws[string.ascii_uppercase[i+1] + str(1)].font = Font(bold=True)
            else:  
                ws[string.ascii_uppercase[i+1] + str(1)] = keywords[i]
                ws[string.ascii_uppercase[i+1] + str(1)].font = Font(bold=True)
        book.save(f'{download_path}\{file_name}')
        book.close()
    
    def write_result(download_path, file_name, result):
        c1 = 1;
        c2 = 1;
        c3 = 1;
        
        book = load_workbook(f'{download_path}\{file_name}')
        ws = book.active
        keyword_1 = ws['B1'].value
        keyword_2 = ws['C1'].value
        keyword_3 = ws['D1'].value
        
        for r in result:
            for i, x in enumerate(r, start=0):
                if keyword_1 in r[i]:
                    c1 += 1
                    ws['B' + str(c1)] = r[i]
                elif keyword_2 in x:
                    c2 += 1
                    ws['C' + str(c2)] = r[i]
                elif keyword_3 in x:
                    c3 += 1
                    ws['D' + str(c3)] = r[i]
        book.save(f'{download_path}\{file_name}')
        book.close()
